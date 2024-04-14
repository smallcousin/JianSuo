import json
import os
import threading
import urllib, urllib3, sys
from datetime import datetime, timedelta
from urllib.parse import urlencode
import ssl

import openpyxl
import pandas as pd

host = 'https://nawater.market.alicloudapi.com'
path = '/api/surface_water/data'
method = 'GET'
appcode = '4431d825fc93415c9d5174836961175a'

# 获取水质数据
def get_query_params(staname, sta_time):
    return {
        'staname': staname,
        'sta_time': sta_time,
        'version': 'v2'
    }


def write_to_excel(data, staname):
    new_column_name =  {
      "province": "身份",
      "city": "城市",
      "river": "河流",
      "basin": "流域",
      "section": "站点",
      "monitor_time": "发布时间",
      "qulity": "水质类别",
      "water_temp": "水温",
      "ph": "PH值",
      "ph_l": "PH等级",
      "dissolvedoxygen": "溶解氧",
      "dissolvedoxygen_l": "溶解氧等级",
      "codmn": "高锰酸盐指数",
      "codmn_l": "溶解氧等级",
      "nh3-n": "氨氮",
      "nh3-n_l": "氨氮等级",
      "tp": "总磷",
      "tn": "总氮",
      "conductivity": "电导率",
      "turbidity": "浊度",
      "chlorophyll": "叶绿素",
      "algal_density": "藻密度",
      "section_status": "站点情况"
    }
    df = pd.DataFrame(data)
    df = df.rename(columns=new_column_name)
    file_path = '..\excel'
    if not os.path.exists(file_path):
        os.makedirs(file_path)
        # 指定文件路径，注意这里路径已经调整为上一级目录
    full_file_name = f'{file_path}\{staname}水质数据.xlsx'
    if not os.path.exists(full_file_name):
        with pd.ExcelWriter(full_file_name, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False)
        print('新建文件-------------' + full_file_name)
    else:
        # with pd.ExcelWriter(full_file_name, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
        #     df.to_excel(writer, index=False, header=not writer.book.sheetnames)
        # 使用openpyxl加载现有的Excel文件
        wb = openpyxl.load_workbook(full_file_name)
        # 选择工作簿中的特定工作表，如果只有一个工作表可以使用active
        sheet = wb.active
        # 找到第一个空白行（即最后一行数据后的下一行）
        row = sheet.max_row + 1
        # 追加数据
        for index, row_data in df.iterrows():
            for col, value in enumerate(row_data, start=1):
                sheet.cell(row=row, column=col, value=value)
            row += 1  # 准备下一行的数据写入

        # 保存工作簿
        wb.save(full_file_name)


def time_plus_one_day(time):
    one_day_later = time + timedelta(days=1)
    return one_day_later


# 写入一条数据
def get_water_data(staname, search_time):
    # 创建PoolManager实例
    http = urllib3.PoolManager()
    # 添加headers
    headers = {
        'Authorization': 'APPCODE ' + appcode
    }
    # 发起请求并读取响应
    fields = get_query_params(staname, search_time)
    print(fields)
    response = http.request('GET', host + path, fields=fields, headers=headers)
    # 输出响应内容
    response_json = json.loads(response.data)
    print(response_json['data'])
    if not response_json['data']:
        print(staname + search_time + '的数据为空')
    else:
        write_to_excel(response_json['data'], staname)
        print(staname + '站点' + search_time + '时间段的数据写入成功')

# 写入一个站点的数据
def get_one_station_water_data(dict):
    start_time = datetime.strptime(dict['start_time'], '%Y-%m-%d %H:%M:%S')
    end_time = datetime.strptime(dict['end_time'], '%Y-%m-%d %H:%M:%S')
    while start_time < end_time:
        get_water_data(dict['staname'], start_time.strftime('%Y-%m-%d %H:%M:%S'))
        start_time = time_plus_one_day(start_time)


if __name__ == '__main__':
    print('======start=======')
    keys = ['staname', 'start_time', 'end_time']
    lakes_values = [
        # ['新丰江水库', '2020-04-30 12:00:00', '2022-06-13 12:00:00'], # 2022-06-13 12:00:00
        # ['良德水库', '2020-04-30 12:00:00', '2024-03-06 04:00:00'], # 2024-03-06 04:00:00
        # ['石骨水库', '2020-04-30 12:00:00', '2024-03-05 20:00:00'] # 2024-03-05 20:00:00
        # ['七星岗', '2020-04-30 12:00:00', '2024-03-06 00:00:00'],
        # ['沙河河口(里波水）', '2020-04-30 12:00:00', '2024-03-06 00:00:00'],
        # ['惠州汝湖', '2020-04-30 12:00:00', '2024-03-06 00:00:00'],
        ['鄱阳湖出口', '2020-04-30 12:00:00', '2021-01-10 08:00:00'],
        ['东洞庭湖', '2020-04-30 12:00:00', '2023-10-13 16:00:00'],
        ['胥湖心', '2020-04-30 12:00:00', '2024-04-06 04:00:00']
    ]
    dicts = [dict(zip(keys, value)) for value in lakes_values]
    # 多线程跑数据
    for d in dicts:
        thread = threading.Thread(target=get_one_station_water_data, kwargs={'dict': d})
        thread.start()
        print('=========================' + d['staname'] + '站点的数据已跑完===========================================')
    # if response.data:
    #     print(response.data.decode('utf-8'))
